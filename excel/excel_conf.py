"""
_*_ coding = utf-8 _*_ 
@time:2020/7/5:11:21
"""

import openpyxl
from log.log import Logger
from openpyxl.styles import PatternFill, Font
import yaml


class Excel_conf:
    # 创建日志对象
    log = Logger().get_logger()

    # 定义Excel格式
    def cell_write(self, value, sheet, row):
        bold = Font(bold=True)
        if value == 'pass':
            fill = PatternFill('solid', fgColor='AACF91')
        elif value == 'false':
            fill = PatternFill('solid', fgColor='FF0000')
        else:
            pass
        sheet.cell(row=row, column=8).value = value.upper()
        sheet.cell(row=row, column=8).fill = fill
        sheet.cell(row=row, column=8).font = bold

    # 读取配置文件中的信息，获取excel路径，基于yaml获取文件
    def load_yaml(self, yaml_path):
        file = open(yaml_path, 'r', encoding='utf-8')
        file_data = yaml.load(file, Loader=yaml.FullLoader)
        return file_data

    # openpyxl 打开excel用例
    def open_excel(self, excel_path):
        # 读取excel的内容
        excel = openpyxl.load_workbook(excel_path)
        return excel

    # 获取excel的所有sheet
    def get_sheets(self, excel):
        sheets = excel.sheetnames
        return sheets

    # 关闭excel文件，释放资源
    def close(self, excel):
        excel.close()

    # 保存excel
    def excel_save(self, excel, path):
        excel.save(path)
