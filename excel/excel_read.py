# __**__ coding=utf-8 __**__
# 作者：calm_zn
# 日期：2020/6/30 17:31
# 工具：PyCharm
# Python版本：3.7
# https://www.imooc.com/article/302606
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet('Mysheet')
ws2 = wb.create_sheet('Mysheet', 0)
ws3 = wb.create_sheet('Mysheet', -1)
ws.title = "New Title"
print(wb.sheetnames)
for sheet in wb:
    print(sheet)


